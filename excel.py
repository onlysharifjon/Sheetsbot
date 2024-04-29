import openpyxl
import os
async def read_excel_part1(path, file_name):
    teachers = []
    all_data = []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    b_values = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
    c_values = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
    sheets_number_rows = 0
    new_list_teachers = []
    new_list_counts = []
    for i in c_values:
        sheets_number_rows += 1
        if i is not None:
            if ',' in str(i).strip():
                soni = str(i).strip().count(',')
                new_list_teachers.append(b_values[sheets_number_rows - 1])
                new_list_counts.append(soni)
    teacher_counts = {}
    for teacher, count in zip(new_list_teachers, new_list_counts):
        if teacher in teacher_counts:
            teacher_counts[teacher] += count
        else:
            teacher_counts[teacher] = count
    for value in b_values:
        all_data.append(str(value))
        if value is not None and str(value).strip() not in teachers:
            a = str(value).strip()
            teachers.append(a)
    fake_database = {}
    txt = ''
    for i in teachers:
        count = 0
        teacher_name = i
        for k in all_data:
            if i in k:
                count += 1
        fake_database[f'{teacher_name}'] = count
    wb.close()
    barcha_qatnashganlar = sum(fake_database.values())
    normalized_databse = {}
    for ism, son in fake_database.items():
        normalized_databse[ism] = son
    for name,count in teacher_counts.items():
        if name.strip() in normalized_databse:
            normalized_databse[name.strip()] +=count
    print(normalized_databse)
    new_data = {}
    for key, value in normalized_databse.items():
        lowercase_key = key.lower()
        if lowercase_key in new_data:
            new_data[lowercase_key] += value
        else:
            new_data[lowercase_key] = value
    cnt = 0
    for name, count in new_data.items():
        cnt+=1
        txt += f"{cnt} 🧑‍🏫 #{name.title()} : <code>{count}</code>\n"
    os.remove(f'uploads/{file_name}')
    return txt
